"""
Pla de comunicació Criteri ESG - agost i setembre 2026
Genera un Excel multi-full amb calendari editorial, peces aliades, KPIs, guia de to i assets visuals.
"""
import sys, os
XLSX_SKILL_DIR = "/home/z/my-project/skills/xlsx"
for sub in [XLSX_SKILL_DIR, os.path.join(XLSX_SKILL_DIR, "templates")]:
    if sub not in sys.path:
        sys.path.insert(0, sub)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import CellIsRule, FormulaRule

# ===== PALETA Criteri ESG (terra + coure, del DESIGN_SYSTEM.md) =====
PRIMARY = "2C1810"       # marró molt fosc - títols
SECONDARY = "5C3A1E"     # marró fosc - capçaleres taula
ACCENT = "B87333"        # coure - accents, hover
ACCENT_LIGHT = "E8C99A"  # coure clar - fills subtils
CREAM = "F5EFE6"         # fons clar
RULE = "C9B89A"          # línies
MUTED = "8B7355"         # text secundari
HOVER = "8A5526"         # accent fosc
WHITE = "FFFFFF"

# Estils compartits
def thin(color=RULE):
    return Side(style="thin", color=color)

def medium(color=SECONDARY):
    return Side(style="medium", color=color)

# Fonts (en Excel no es poden carregar Fraunces/Inter directament; uso fonts del sistema amb mateix esperit)
FONT_TITLE = "Georgia"       # serif per títols (similar a Fraunces)
FONT_BODY = "Calibri"        # sans per body (similar a Inter)
FONT_MONO = "Consolas"       # mono per meta (similar a JetBrains Mono)

def style_title(cell):
    cell.font = Font(name=FONT_TITLE, size=20, bold=True, color=PRIMARY)
    cell.alignment = Alignment(horizontal="left", vertical="center")

def style_subtitle(cell):
    cell.font = Font(name=FONT_BODY, size=11, italic=True, color=MUTED)
    cell.alignment = Alignment(horizontal="left", vertical="center")

def style_section_header(cell):
    cell.font = Font(name=FONT_TITLE, size=14, bold=True, color=PRIMARY)
    cell.alignment = Alignment(horizontal="left", vertical="center")

def style_table_header(cell):
    cell.font = Font(name=FONT_BODY, size=10, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=SECONDARY)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = Border(bottom=medium(ACCENT))

def style_data_cell(cell, wrap=True, bold=False, color=PRIMARY, size=10, fill=None):
    cell.font = Font(name=FONT_BODY, size=size, bold=bold, color=color)
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=wrap)
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)

def style_meta(cell):
    cell.font = Font(name=FONT_MONO, size=9, color=MUTED)
    cell.alignment = Alignment(horizontal="left", vertical="center")

# ===== WORKBOOK =====
wb = Workbook()
wb.properties.creator = "Z.ai · Criteri ESG"
wb.properties.title = "Pla de comunicació agost-setembre 2026"

# Elimina el full per defecte al final
default = wb.active

# ============================================================
# FULL 1 — RESUM EXECUTIU
# ============================================================
ws1 = wb.create_sheet("Resum executiu")
ws1.sheet_view.showGridLines = False
ws1.column_dimensions['A'].width = 2
ws1.column_dimensions['B'].width = 22
ws1.column_dimensions['C'].width = 80

# Capçalera de marca
ws1['B2'] = "Criteri ESG"
style_title(ws1['B2'])
ws1['B3'] = "Pla de comunicació · agosetembre 2026 · versió 1.0"
style_subtitle(ws1['B3'])
ws1.row_dimensions[2].height = 30

# Linia accent
ws1['B4'] = ""
ws1['B4'].fill = PatternFill("solid", fgColor=ACCENT)
ws1.row_dimensions[4].height = 4
ws1.merge_cells('B4:C4')

# === Bloc 1: Objectius ===
ws1['B6'] = "Objectius"
style_section_header(ws1['B6'])
ws1.row_dimensions[6].height = 24

objectius = [
    ("Agost (sembrar)", "Construir presència editorial sense anunciar producte. La gent ha de descobrir Criteri ESG com 'algú que finalment diu en veu alta allò que molts pensem'. Objectiu: 80-120 subscriptors a la waitlist, 200-400 seguidors LinkedIn empresa, 4 peces aliades publicades."),
    ("Setmana 1-2 set (soft launch)", "Anunciar el projecte sense estirabots. Mostrar la web com a 'ja és aquí', no com a 'arriba aviat'. Obertura del registre. Objectiu: 300-500 visitants únics, 50-80 registres."),
    ("Setmana 3-5 set (consolidació)", "Publicar el contingut de fons (articles, metodologia, preguntes ètiques) que justifica el valor del producte. La gent prova, llegeix, decideix. Objectiu: 800-1.500 visitants, 100-150 registres, 10-15 converses iniciades."),
]
row = 8
for label, txt in objectius:
    ws1.cell(row=row, column=2, value=label)
    style_data_cell(ws1.cell(row=row, column=2), wrap=True, bold=True, color=HOVER, size=10)
    ws1.cell(row=row, column=3, value=txt)
    style_data_cell(ws1.cell(row=row, column=3), wrap=True, size=10)
    ws1.row_dimensions[row].height = 60
    row += 1

# === Bloc 2: Posicionament ===
row += 1
ws1.cell(row=row, column=2, value="Posicionament")
style_section_header(ws1.cell(row=row, column=2))
ws1.row_dimensions[row].height = 24
row += 1
ws1.cell(row=row, column=2, value="Una frase")
style_data_cell(ws1.cell(row=row, column=2), bold=True, color=HOVER, size=10)
ws1.cell(row=row, column=3, value="Criteri ESG no és un agregador d'informes. És un filtre amb criteri: què diu, què et afecta, què has de fer.")
style_data_cell(ws1.cell(row=row, column=3), wrap=True, size=11)
ws1.row_dimensions[row].height = 36
row += 1
ws1.cell(row=row, column=2, value="Tres verbs")
style_data_cell(ws1.cell(row=row, column=2), bold=True, color=HOVER, size=10)
ws1.cell(row=row, column=3, value="Sintetitzar (informes), Creuar (amb les certificacions que ja tens), Ajudar a decidir (accions recomanades). Res més.")
style_data_cell(ws1.cell(row=row, column=3), wrap=True, size=11)
ws1.row_dimensions[row].height = 36
row += 1
ws1.cell(row=row, column=2, value="Què NO som")
style_data_cell(ws1.cell(row=row, column=2), bold=True, color=HOVER, size=10)
ws1.cell(row=row, column=3, value="No som una newsletter més de notícies ESG. No som un agregador. No som una consultoria. No substituïm el director de sostenibilitat: li estalviem temps perquè pugui pensar.")
style_data_cell(ws1.cell(row=row, column=3), wrap=True, size=11)
ws1.row_dimensions[row].height = 48

# === Bloc 3: Principis de to ===
row += 2
ws1.cell(row=row, column=2, value="Principis de to")
style_section_header(ws1.cell(row=row, column=2))
ws1.row_dimensions[row].height = 24
row += 1
principis = [
    ("1. Editorial, no comercial", "Res de 'apunta't ara', 'no te'l perdis', 'oferta limitada'. El nostre to és el d'un observatori que publica, no el d'una marca que ven."),
    ("2. Una dada per peça", "Cada peça té una sola idea amb una sola dada. La dada sempre és real, sempre està referenciada. Mai inventada."),
    ("3. Sempre una pregunta", "Cada peça acaba amb una pregunta oberta, no amb un CTA. La pregunta crea relació; el CTA crea transacció."),
    ("4. Mai superlatius", "Prohibits: 'revolucionari', 'innovador', 'únic', 'el primer', 'el millor'. Si ho som, que ho diguin els altres."),
    ("5. Referenciar sempre", "Cada afirmació sobre un informe o regulació porta la font. Sempre. És el nostre actiu de credibilitat."),
]
for label, txt in principis:
    ws1.cell(row=row, column=2, value=label)
    style_data_cell(ws1.cell(row=row, column=2), wrap=True, bold=True, color=HOVER, size=10)
    ws1.cell(row=row, column=3, value=txt)
    style_data_cell(ws1.cell(row=row, column=3), wrap=True, size=10)
    ws1.row_dimensions[row].height = 42
    row += 1

# === Bloc 4: Peces aliades ===
row += 1
ws1.cell(row=row, column=2, value="Peces aliades")
style_section_header(ws1.cell(row=row, column=2))
ws1.row_dimensions[row].height = 24
row += 1
aliades = [
    ("Mapa dels 16 estàndards ESG", "Visual de referència amb els 16 estàndards organitzats en 3 categories (Regulacions, Frameworks, Certificacions) amb colors terra/coure. Llançament: 12 agost. Reutilització: cada vegada que esmenti un estàndard, s'enllaça al mapa."),
    ("Llista de les 30 fonts que monitoritem", "Document descarregable PDF amb les 30 fonts més importants, cadència de publicació i per què importen. Llançament: 19 agost. Reutilització: actualització mensual, menció a cada newsletter."),
    ("Carta ètica 'Més enllà del Checkbox'", "Sèrie de 5 preguntes ètiques publicades cada quinzena. Primera: 'Si tu empresa desapareciera mañana...'. Llançament: 26 agost. Reutilització: una peça per quinzena durant tot el curs."),
    ("Semàfor metodològic", "Visual que explica com avaluem els informes (5 dimensions: Scope 3, Plazos, Fuentes, Granularidad, Verificación). Llançament: 23 setembre. Reutilització: cada informe porta el seu semàfor."),
]
for label, txt in aliades:
    ws1.cell(row=row, column=2, value=label)
    style_data_cell(ws1.cell(row=row, column=2), wrap=True, bold=True, color=HOVER, size=10)
    ws1.cell(row=row, column=3, value=txt)
    style_data_cell(ws1.cell(row=row, column=3), wrap=True, size=10)
    ws1.row_dimensions[row].height = 60
    row += 1

# === Bloc 5: Canals ===
row += 1
ws1.cell(row=row, column=2, value="Canals i responsabilitats")
style_section_header(ws1.cell(row=row, column=2))
ws1.row_dimensions[row].height = 24
row += 1
canals = [
    ("LinkedIn empresa (principal)", "Publicació dimarts i dijous. To editorial. Una veu col·lectiva ('Criteri ESG'), no personal. Objectiu: 200-400 seguidors agost, 600-1.000 setembre."),
    ("Newsletter (Beehiiv)", "Bimensual a partir de 3 setembre (dijous 15:00h). Dues versions: gratis (resum + connexió) i completa (Premium). Objectiu: 80 subs agost, 250 subs finals setembre."),
    ("Web / blog (SEO)", "Articles llargs publicats al blog. Cadascun és una peça per si mateix i alimenta LinkedIn. Objectiu: 4 articles en 2 mesos."),
    ("Twitter/X (complementari)", "Republishing de peces de LinkedIn, sense esforç addicional. Sense objectius."),
    ("Personal Paolo (mínim)", "3 posts personals en 2 mesos, només en moments clau (presentació, llançament, balanç). To fundador, no tècnic."),
]
for label, txt in canals:
    ws1.cell(row=row, column=2, value=label)
    style_data_cell(ws1.cell(row=row, column=2), wrap=True, bold=True, color=HOVER, size=10)
    ws1.cell(row=row, column=3, value=txt)
    style_data_cell(ws1.cell(row=row, column=3), wrap=True, size=10)
    ws1.row_dimensions[row].height = 48
    row += 1

# Footer
row += 2
ws1.cell(row=row, column=2, value="Document viu · actualitzat 19 juliol 2026 · Z.ai per a Criteri ESG")
style_meta(ws1.cell(row=row, column=2))
ws1.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)

ws1.sheet_view.zoomScale = 110

# ============================================================
# FULL 2 — CALENDARI EDITORIAL (la peça principal)
# ============================================================
ws2 = wb.create_sheet("Calendari editorial")
ws2.sheet_view.showGridLines = False

# Títol
ws2['B2'] = "Calendari editorial · agost-setembre 2026"
style_title(ws2['B2'])
ws2.row_dimensions[2].height = 28
ws2['B3'] = "Una fila per peça de comunicació. Textos complets en castellà. Filtra per Setmana o Canal."
style_subtitle(ws2['B3'])

# Linia accent
ws2['B4'] = ""
ws2['B4'].fill = PatternFill("solid", fgColor=ACCENT)
ws2.row_dimensions[4].height = 3
ws2.merge_cells('B4:L4')

# Capçaleres
headers = [
    "Setmana", "Data", "Fase", "Canal", "Format",
    "Títol / temàtica", "Text complet", "Asset visual suggerit",
    "CTA", "KPI principal", "Responsable"
]
col_widths = [12, 12, 16, 18, 18, 28, 70, 32, 28, 22, 14]
for i, (h, w) in enumerate(zip(headers, col_widths), start=2):
    col_letter = get_column_letter(i)
    ws2.column_dimensions[col_letter].width = w
    cell = ws2.cell(row=6, column=i, value=h)
    style_table_header(cell)
ws2.row_dimensions[6].height = 32

# === LES 18 PECES ===
# Format: (setmana, data, fase, canal, format, titol, text, asset, cta, kpi, responsable)
peces = [
    # ===== AGOST - SEMBRAR =====
    ("W1·agost", "dimarts 5 ago", "Sembrar", "LinkedIn empresa", "Post curt (6 línies)",
     "El soroll ESG",
     "Cada semana salen 14 informes institucionales sobre sostenibilidad. Todos importantes. Todos largos. Todos con datos que cabría leer.\n\nLa pregunta no es qué publicará la UE mañana. La pregunta es: quién te ayuda a saber qué hacer con aquello que ya se ha publicado y aún no has tenido tiempo de abrir.\n\nEstamos preparando algo.",
     "Imagen minimalista fondo crema #F5EFE6 con tipografía Fraunces '14 informes. 0 tiempo.' Centrada. Sin logo.",
     "Ninguno explícito. Al final: 'Pronto.'",
     "Impresiones + comentarios",
     "Paolo (publica)"),

    ("W1·agost", "dijous 7 ago", "Sembrar", "LinkedIn empresa", "Post curt (8 línies)",
     "Què fa un director de sostenibilitat entre març i juny",
     "Está leyendo la revisión de los ESRS. Está intentando entender el cambio al CSDDD. Está ajustando el B Impact Assessment porque B Lab ha cambiado la ponderación. Está recibiendo presión de inversores para subir el rating MSCI.\n\nY tiene 47 páginas de la Comisión Europea en la bandeja de entrada.\n\nNo le falta información. Le falta tiempo para procesarla. Y criterio para decidir qué importa.",
     "Foto en B/N de una mesa con papeles, o gráfico minimalista con 4 tareas superpuestas",
     "Ninguno",
     "Impresiones + shares",
     "Paolo"),

    ("W2·agost", "dimarts 12 ago", "Sembrar", "LinkedIn empresa + Web", "Article llarg (700 paraules) + post que l'anuncia",
     "El mapa dels 16 estàndards ESG que cap director hauria de confondre",
     "POST LINKEDIN (anunci):\nEsta semana publicamos algo que llevamos meses preparando: el mapa de los 16 estándares ESG con los que trabaja un director de sostenibilidad español.\n\nLos distinguimos en tres categorías con tres colores:\n— Regulaciones (CSRD, CSDDD, SFDR, Taxonomía UE, EMAS): te obligan.\n— Frameworks (GRI, SASB, TNFD, TCFD, ISO 26000): te orientan.\n— Certificaciones y ratings (EcoVadis, B Corp, MSCI, CDP, SGE 21, Sustainalytics): te evalúan.\n\nConfundirlas tiene consecuencias operativas. No las confundáis.\n\nMapa completo: criteriesg.com/estandares-esg\n\nARTICLE WEB (esborrany per completar):\n[Article de 700 paraules explicant: per què els vam classificar així, què canvia segons la categoria, com es creuen entre ells, exemples reals de quan la confusió genera errors operatius. Acaba amb: 'Aquest mapa és viu: l'actualitzarem cada trimestre'.]",
     "Mapa visual dels 16 estàndards amb 3 colors (Regulacions #5C3A1E, Frameworks #B87333, Certificacions #E8C99A). Format quadrat 1080x1080 per LinkedIn + versió horitzontal 1200x630 per Open Graph.",
     "criteriesg.com/estandares-esg",
     "Clics a la web + saves LinkedIn",
     "Paolo + Z.ai (text)"),

    ("W2·agost", "dijous 14 ago", "Sembrar", "LinkedIn empresa", "Post curt (7 línies)",
     "El malentès comú: CSRD no és una certificació",
     "Esta semana publicamos nuestro mapa de estándares ESG. La reacción más frecuente: 'no sabía que CSRD era una regulación, no una certificación'.\n\nLa distinción importa.\nUna certificación la pides. Una regulación te obliga.\nOtra cosa es el framework con el que la recibes (GRI, ESRS). Otra el rating que te darán (MSCI, Sustainalytics).\n\nConfundirlas tiene consecuencias operativas. No las confundáis.",
     "Detalle del mapa con CSRD destacado en su categoría 'Regulación'",
     "criteriesg.com/estandares-esg",
     "Clics + comentarios",
     "Paolo"),

    ("W3·agost", "dimarts 19 ago", "Sembrar", "LinkedIn empresa + Newsletter interna", "Article curt (400 paraules) + email a la waitlist",
     "Les 30 fonts que estem llegint (i per què)",
     "POST LINKEDIN:\nCada lunes y jueves a las 9:00 revisamos 192 fuentes. Banco Central Europeo, ESRS Q&A Platform, EcoVadis Methodology updates, Sustainalytics Risk Reports, Forética, B Lab Spain, MSCI ESG Ratings, CDP, Global Reporting Initiative, TNFD...\n\nHoy compartimos las 30 que consideramos imprescindibles para un director de sostenibilidad español. No es una lista de 'nice to have'. Es nuestra rutina diaria.\n\nUsadla.\n\nEMAIL WAITLIST:\nHola,\nHace una semana te apuntaste a Criteri ESG. Hoy te compartimos algo concreto: las 30 fuentes que leemos cada semana.\nEs nuestra forma de decirte qué hacemos cuando no nos ves. Es también nuestra forma de demostrarte que no improvisamos.\nEn dos semanas: el primer número de nuestra newsletter.\n— Paolo",
     "Tabla visual con 30 logos organizados por tipología (Reguladores / ONGs / Ratings / Acadèmia). Fons crema.",
     "criteriesg.com/fuentes (pàgina nova) + 'Apúntate a la newsletter'",
     "Subscriptors waitlist + clics",
     "Paolo + Z.ai"),

    ("W3·agost", "dijous 21 ago", "Sembrar", "LinkedIn empresa", "Post reflexiu (10 línies)",
     "El silenci dels que ja ho sabien",
     "Hay una frase que escuchamos a menudo en conversaciones con directores de sostenibilidad: 'esto ya lo sabía, pero no tenía manera de demostrarlo'.\n\nEl problema del ESG no es la falta de información. Es la falta de tiempo para procesarla y de criterio para decir qué importa.\n\nY este 'saber sin poder demostrar' tiene un coste: decisiones que se toman con intuición en lugar de datos, ratificadas después por inercia.\n\nNuestro trabajo es convertir ese saber implícito en saber demostrable. Nada más.",
     "Imagen minimalista con la frase 'saber sin poder demostrar' sobre fondo marró #5C3A1E",
     "Ninguno",
     "Impresiones + comentarios cualitativos",
     "Paolo"),

    ("W4·agost", "dimarts 26 ago", "Sembrar", "LinkedIn empresa + Web", "Post + peça visual",
     "Més enllà del Checkbox #1",
     "Si tu empresa desapareciera mañana, ¿quién lo notaría de verdad —y por qué?\n\nLa respuesta te dice más sobre tu valor real que cualquier métrica ESG. Lo escribimos en nuestro manifiesto: las empresas que solo crean valor para accionistas no dejan vacío cuando desaparecen. Las que crean valor para trabajadores, comunidad y territorio, sí.\n\nPiensa en qué vacío dejarías.\n\nCada quincena, una pregunta distinta.",
     "Tipografía Fraunces sobre fondo marró #2C1810 con la pregunta en color crema #F5EFE6. Formato 1080x1350.",
     "criteriesg.com/manifest (pàgina nova)",
     "Impresiones + comentarios (mètrica qualitativa)",
     "Paolo + Z.ai (text)"),

    ("W4·agost", "dijous 28 ago", "Sembrar", "LinkedIn empresa", "Post curt (5 línies)",
     "Cinc setmanes",
     "Cinco semanas. Cinco piezas.\n\nMapa de estándares. Fuentes que leemos. Preguntas que nos hacemos.\n\nA poco a poco. Sin prisa, sin ruido.\n\nSi quieres ser de los primeros en recibir lo que iremos publicando, apúntate. No te enviaremos nada comercial. Solo nuestro trabajo.",
     "Captura de pantalla de la web con countdown sutil '5 semanas' en una cantonada. Sense grans tipografies.",
     "criteriesg.com · 'Apúntate a la newsletter'",
     "Subscriptors waitlist",
     "Paolo"),

    # ===== SETEMBRE - LLANÇAMENT =====
    ("W1·setembre", "dimarts 1 set", "Llançar", "LinkedIn empresa + Newsletter oficial #1 + Personal Paolo", "Post fundacional + newsletter",
     "Criteri ESG, en marxa",
     "POST LINKEDIN EMPRESA:\nHoy lanzamos Criteri ESG.\n\nUna plataforma que sintetiza informes institucionales ESG y los cruza con las certificaciones que tu empresa ya tiene (o debería tener).\n\nNo es un agregador. Es un filtro con criterio: qué dice el informe, qué te afecta, qué has de hacer.\n\nPuedes empezar a leer en criteriesg.com.\nLa primera newsletter sale el jueves a las 15:00h.\n\n—\nPOST PERSONAL PAOLO:\nHace 8 meses dejé mi trabajo para construir esto. Lo hago porque durante años fui el director de sostenibilidad que recibía 14 informes por semana y no tenía tiempo de leerlos. Y porque la solución no era más información: era criterio.\n\nHoy lanzo Criteri ESG. No os pido que os suscribáis. Os pido que lo leáis y me digáis qué falla.\n\n— Paolo",
     "Captura de la homepage o hero image. Sin mockups comerciales.",
     "criteriesg.com · 'Septiembre: acceso abierto a todo el contenido'",
     "Visitants web + registres waitlist",
     "Paolo"),

    ("W1·setembre", "dijous 3 set", "Llançar", "Newsletter", "Newsletter oficial #1",
     "Edició #001 · 'El CSRD se simplifica, pero el CSDDD se endurece'",
     "Newsletter completa amb: header + carta del director + 3 informes destacats (ESRS revisió, CSDDD Omnibus I, EcoVadis Q1) + connexió editorial entre ells + notícies ESG (3 titulars) + inversió ESG + pregunta ètica.\n\nFormat: com el mockup validat (sense CTAs Premium, semàfor compacte a dalt a la dreta).\n\nL'enllaç principal: 'Más detalles en el informe completo'.",
     "Cap asset extern. La newsletter és la peça.",
     "criteriesg.com/informes/[slug] per cada informe",
     "Open rate >40%, CTR >10%",
     "Z.ai (text) + Paolo (envia)"),

    ("W2·setembre", "dimarts 9 set", "Consolidar", "LinkedIn + Article web", "Article llarg (1.500 paraules)",
     "Per què els directors de sostenibilitat perden el 60% del seu temps recopilant informació",
     "Un director de sostenibilidad dedica, de media, el 60% de su jornada a recopilar, leer y estructurar información procedente de fuentes externas: reguladores, frameworks, ratings, certificadores. Solo el 15% lo dedica a analizar. Solo el 10%, a decidir.\n\n[Article desenvolupant: dades de l'estudi (referenciar), 3 casos anonimitzats de directors reals, la paradoxa del reporting ESG (més informació = menys decisió), i la proposta de Criteri: filtrar, sintetitzar, creuar. Sense dir 'compreu Criteri'.]\n\nAcaba amb: 'Si reconoces esta situación, te entendemos. Estamos aquí para devolverte el 60%.'",
     "Gràfic visual 'On es perd el temps del director ESG' - barres horitzontals amb 4 categories (Recopilar 60%, Llegir 15%, Analitzar 15%, Decidir 10%). Colors terra.",
     "criteriesg.com · 'Lee este artículo completo + apúntate a la newsletter'",
     "Lectures article + registres",
     "Z.ai (text) + Paolo (publica)"),

    ("W2·setembre", "dijous 11 set", "Consolidar", "LinkedIn empresa", "Post curt (8 línies)",
     "El que no ens diu l'informe",
     "Esta semana hemos procesado 4 informes. Todos rigurosos. Todos útiles.\n\nPero todos dejan fuera la misma cosa: qué has de hacer tú.\n\nCuando un director de sostenibilidad cierra un informe, su pregunta no es '¿qué dice?' sino '¿y ahora qué?'. Y ese 'y ahora qué' depende de qué certificaciones tienes.\n\nPor eso hacemos cross-reference con EcoVadis, B Corp, MSCI, GRI, SGE 21...\n\nLa dato sin tu contexto es ruido. La dato con tu contexto es decisión.",
     "Captura d'un informe amb la secció cross-reference destacada (versió retallada)",
     "criteriesg.com/informes · 'Lee los 4 informes'",
     "Clics + registres",
     "Paolo"),

    ("W3·setembre", "dimarts 15 set", "Consolidar", "LinkedIn + Web", "Peça reflexiva (500 paraules)",
     "Més enllà del Checkbox #2",
     "Pregunta de esta quincena:\n\nSi mañana te pidieran demostrar el impacto positivo de tu empresa con un solo dato, ¿cuál elegirías y por qué?\n\nLa mayoría de directores no sabe responder. Tienen 47 KPIs y ninguno que sintetice. Tener demasiadas métricas sin un relato coherente es tan inútil como no tener ninguna.\n\nLa cuestión no es cuantificar más, es saber qué cuenta.\n\n[Article curt desenvolupant la idea amb 2 exemples reals anonimitzats i acabant amb una segona pregunta: 'I si no saps quin dada triar, és la dada equivocada o el relat equivocat?']",
     "Tipografia Fraunces sobre fons crema #F5EFE6 amb la pregunta en color marró #2C1810. Format 1080x1350.",
     "criteriesg.com/manifest · 'Cada quincena, una pregunta'",
     "Comentaris (mètrica qualitativa)",
     "Paolo + Z.ai (text)"),

    ("W3·setembre", "dijous 17 set", "Consolidar", "Newsletter", "Newsletter oficial #2",
     "Edició #002 · 'EcoVadis canvia el joc per als reportings GRI'",
     "Newsletter #2. Mates estructura que #1. 3 informes nous del període 4-17 setembre + connexió editorial + notícies ESG + pregunta ètica #3.",
     "Cap asset extern.",
     "criteriesg.com/informes/[slug]",
     "Open rate >45%, CTR >12%",
     "Z.ai (text) + Paolo (envia)"),

    ("W4·setembre", "dimarts 23 set", "Consolidar", "LinkedIn + Article web", "Article mig (800 paraules)",
     "El semàfor metodològic: com avaluem el que llegim",
     "No todos los informes se crean igual. Un informe de 80 páginas del BCE no es comparable con un press release de 3 páginas de MSCI. Y aun así, los directores los reciben en la misma bandeja de entrada.\n\nPor eso hemos creado un semáforo metodológico con cinco dimensiones:\n\n— Scope 3: ¿cubre emisiones indirectas de la cadena de valor?\n— Plazos: ¿las fechas son operativas (con calendario) o aspiracionales?\n— Fuentes: ¿están referenciadas o son afirmaciones sin respaldo?\n— Granularidad: ¿hay dato desagregado por sector/geografía/tamaño?\n— Verificación: ¿está auditado externamente o es auto-reportado?\n\n[Article desenvolupant cada dimensió amb un exemple real, mostrant com un informe concret rep 'C · Débil' en una i 'B · Robusto' en una altra. Acaba amb: 'Cada informe que publiquem porta su semáforo. Es nuestra forma de no confundir rigor con volumen.']",
     "Visual del semáforo con 5 dimensiones + nota final (estil del widget del mockup newsletter, però en versió gran)",
     "criteriesg.com/informes · 'Lee un informe con su semáforo'",
     "Lectures + registres",
     "Z.ai (text) + Paolo (publica)"),

    ("W4·setembre", "dijous 25 set", "Consolidar", "LinkedIn empresa", "Post curt (8 línies)",
     "L'informe que sí hauries d'haver llegit aquest mes",
     "Este mes la Comisión Europea ha publicado la revisión de los ESRS. 47 páginas.\n\nTe hemos hecho un resumen en 7 minutos con nota C · Débil en nuestro semáforo metodológico.\n\n¿Por qué C? Porque simplifica mucho, pero deja fuera el detalle granular que las empresas necesitan para implementar.\n\nSi tienes EcoVadis Plata, te afecta de una manera. Si tienes B Corp, de otra. Eso es lo que hacemos: cruzar el informe con tus certificaciones.",
     "Captura de l'informe amb cross-reference visible (versió retallada, sense dades sensibles)",
     "criteriesg.com/informes/esrs-revision-mayo-2026",
     "Clics + registres",
     "Paolo"),

    ("W5·setembre", "dimarts 30 set", "Consolidar", "LinkedIn empresa + Newsletter interna", "Post balanç + newsletter curta",
     "Un mes després",
     "Hace 30 días lanzamos Criteri ESG. No hemos vendido nada.\n\nHemos publicado 16 informes, 2 newsletters, 4 artículos. Hablamos con [X] directores de sostenibilidad.\n\nAprendizajes:\n1) El problema no es información, es criterio.\n2) Todo el mundo tiene EcoVadis, nadie lo lee igual.\n3) El CSRD sigue siendo una incógnita para el 80% de los directores.\n\nPróxima parada: cada quincena, una conexión entre informes que no encontraréis en ningún otro sitio.\n\nGracias a los que ya estáis. A los que aún no: os esperamos.",
     "Visual amb 3 dades clau (16 informes, 2 newsletters, X converses) en format minimalista terra+coure",
     "criteriesg.com",
     "Tot",
     "Paolo"),
]

# Escriure les peces
row = 7
for p in peces:
    for col_idx, val in enumerate(p, start=2):
        cell = ws2.cell(row=row, column=col_idx, value=val)
        # Estil segons columna
        if col_idx in (2, 3, 4, 5, 6, 11):  # meta, curts
            style_data_cell(cell, wrap=True, size=9)
        elif col_idx == 7:  # títol
            style_data_cell(cell, wrap=True, size=10, bold=True, color=PRIMARY)
        elif col_idx == 8:  # text complet
            style_data_cell(cell, wrap=True, size=10, color="2C1810")
        else:
            style_data_cell(cell, wrap=True, size=9, color=MUTED)
        # Alternar fila segons fase
        if p[2] == "Sembrar":
            cell.fill = PatternFill("solid", fgColor=CREAM)
        elif p[2] == "Llançar":
            cell.fill = PatternFill("solid", fgColor="EFE3D2")  # coure molt clar
        else:  # Consolidar
            cell.fill = PatternFill("solid", fgColor="F0E8DC")
    # Altura segons llargada del text
    text_len = len(p[6])
    if text_len > 500:
        ws2.row_dimensions[row].height = 180
    elif text_len > 300:
        ws2.row_dimensions[row].height = 130
    elif text_len > 150:
        ws2.row_dimensions[row].height = 90
    else:
        ws2.row_dimensions[row].height = 60
    row += 1

# Congelar capçalera
ws2.freeze_panes = "C7"
ws2.sheet_view.zoomScale = 90

# AutoFiltre
ws2.auto_filter.ref = f"B6:L{row-1}"

# ============================================================
# FULL 3 — PECES ALIADES
# ============================================================
ws3 = wb.create_sheet("Peces aliades")
ws3.sheet_view.showGridLines = False

ws3['B2'] = "Peces aliades · assets que sostenen la comunicació"
style_title(ws3['B2'])
ws3.row_dimensions[2].height = 28
ws3['B3'] = "Quatre peces que es comparteixen, es descarreguen, generen SEO i creen diferenciació."
style_subtitle(ws3['B3'])

ws3['B4'] = ""
ws3['B4'].fill = PatternFill("solid", fgColor=ACCENT)
ws3.row_dimensions[4].height = 3
ws3.merge_cells('B4:H4')

headers3 = ["Peça", "Descripció", "Format", "Data llançament", "Canals", "Mètrica èxit", "Reutilització futura"]
widths3 = [28, 60, 22, 18, 28, 28, 38]
for i, (h, w) in enumerate(zip(headers3, widths3), start=2):
    col_letter = get_column_letter(i)
    ws3.column_dimensions[col_letter].width = w
    cell = ws3.cell(row=6, column=i, value=h)
    style_table_header(cell)
ws3.row_dimensions[6].height = 32

aliades_data = [
    ("Mapa dels 16 estàndards ESG",
     "Visual de referència amb els 16 estàndards organitzats en 3 categories (Regulacions, Frameworks, Certificacions) amb colors terra/coure. Cada estàndard porta descripció curta, nombre d'informes cross-ref i tipus d'accés (Gratis/Premium).",
     "Pàgina web interactiu + PDF descarregable + imatge LinkedIn 1080x1080",
     "12 agost 2026",
     "LinkedIn empresa (post llançament + 3 posts derivats), web, newsletter #1",
     "500+ visualitzacions pàgina web, 50+ descarregues PDF, 100+ saves LinkedIn",
     "Permanent. Cada vegada que mencionem un estàndard en un informe, enllacem al mapa. Actualització trimestral."),

    ("Llista de les 30 fonts que monitoritem",
     "Document PDF amb 30 fonts organitzades per tipologia (Reguladors europeus, Reguladors espanyols, Frameworks, Ratings, ONGs, Acadèmia). Per cada font: nom, URL, cadència de publicació, per què importa, què en traiem.",
     "PDF descarregable + pàgina web",
     "19 agost 2026",
     "LinkedIn empresa (post + article curt), email a la waitlist, newsletter #1",
     "80+ subscriptors waitlist originats per la peça, 200+ descarregues PDF",
     "Actualització mensual. Cada mes afegim o eliminem fonts. La versió actual sempre és pública."),

    ("Carta ètica 'Més enllà del Checkbox'",
     "Sèrie de preguntes ètiques publicades cada quinzena. Primera pregunta: 'Si tu empresa desapareciera mañana, ¿quién lo notaría de verdad?'. Cada pregunta va acompanyada d'un text curt de reflexió (300-500 paraules). 5 preguntes planificades per setembre-desembre.",
     "Post LinkedIn + peça visual 1080x1350 + pàgina web acumulativa",
     "26 agost 2026 (Pregunta #1)",
     "LinkedIn empresa (cadascuna), web (pàgina acumulativa), newsletter (apartat 'Pregunta para Mejorar')",
     "10+ comentaris per pregunta (qualitat > quantitat)",
     "Cada quinzena durant tot el curs. Es converteix en una secció fixa de la newsletter i en un actiu de marca."),

    ("Semàfor metodològic",
     "Visual que explica com avaluem els informes en 5 dimensions: Scope 3, Plazos, Fuentes, Granularidad, Verificación. Cada informe té una nota final (A · Robusto fort a D · Insuficiente). Inclou guia visual dels colors.",
     "Article web + visual 1200x630 + widget incrustat a cada informe",
     "23 setembre 2026",
     "LinkedIn empresa (article), web (pàgina dedicada + widget a cada informe), newsletter #3",
     "200+ lectures article, widget present a tots els informes",
     "Permanent. Tots els informes des de setembre porten el seu semàfor. Es converteix en signatura de marca."),
]

row = 7
for a in aliades_data:
    for col_idx, val in enumerate(a, start=2):
        cell = ws3.cell(row=row, column=col_idx, value=val)
        if col_idx == 2:
            style_data_cell(cell, wrap=True, size=11, bold=True, color=PRIMARY)
        elif col_idx == 3:
            style_data_cell(cell, wrap=True, size=10)
        else:
            style_data_cell(cell, wrap=True, size=9, color=MUTED)
        cell.fill = PatternFill("solid", fgColor=CREAM if row % 2 else WHITE)
    ws3.row_dimensions[row].height = 130
    row += 1

ws3.freeze_panes = "C7"
ws3.sheet_view.zoomScale = 100

# ============================================================
# FULL 4 — KPIs I SEGÜIMENT
# ============================================================
ws4 = wb.create_sheet("KPIs i seguiment")
ws4.sheet_view.showGridLines = False

ws4['B2'] = "KPIs i seguiment setmanal"
style_title(ws4['B2'])
ws4.row_dimensions[2].height = 28
ws4['B3'] = "Objectius orientatius (no agressius). Paolo omple 'Real' cada diumenge al vespre. Columna 'Delta' calculada automàticament."
style_subtitle(ws4['B3'])

ws4['B4'] = ""
ws4['B4'].fill = PatternFill("solid", fgColor=ACCENT)
ws4.row_dimensions[4].height = 3
ws4.merge_cells('B4:J4')

headers4 = ["Setmana", "Fase", "Subscriptors newsletter (objectiu)", "Subscriptors newsletter (real)", "Visitants web (objectiu)", "Visitants web (real)", "Seguidors LinkedIn (objectiu)", "Seguidors LinkedIn (real)", "Notes qualitatives"]
widths4 = [14, 14, 18, 18, 16, 16, 18, 18, 50]
for i, (h, w) in enumerate(zip(headers4, widths4), start=2):
    col_letter = get_column_letter(i)
    ws4.column_dimensions[col_letter].width = w
    cell = ws4.cell(row=6, column=i, value=h)
    style_table_header(cell)
ws4.row_dimensions[6].height = 42

kpi_data = [
    ("W1·agost", "Sembrar", 5, None, 50, None, 30, None, "Primera setmana. Sensació de començar de zero. Apuntar qualsevol comentari positiu, encara que sigui 1."),
    ("W2·agost", "Sembrar", 15, None, 100, None, 70, None, "Llançament del mapa d'estàndards. Hauria de ser el primer pic de tràfic."),
    ("W3·agost", "Sembrar", 35, None, 200, None, 130, None, "Llista de 30 fonts. Si la gent la descarrega, tenim un senyal que la transparència funciona."),
    ("W4·agost", "Sembrar", 80, None, 350, None, 200, None, "Pregunta ètica + anunci waitlist. Hauria d'accelerar subscripcions."),
    ("W1·setembre", "Llançar", 150, None, 600, None, 350, None, "Llançament oficial. Tot el que hem sembrat es converteix en tràfic. Pic màxim esperat."),
    ("W2·setembre", "Llançar", 220, None, 850, None, 480, None, "Article profund. El públic més implicat el llegirà."),
    ("W3·setembre", "Consolidar", 280, None, 1.050, None, 580, None, "Newsletter #2. Si l'open rate >45%, tenim un senyal de fidelització."),
    ("W4·setembre", "Consolidar", 320, None, 1.200, None, 660, None, "Semàfor metodològic. Peça que hauria de generar 'ara t'entenc'."),
    ("W5·setembre", "Consolidar", 350, None, 1.400, None, 720, None, "Balanç. Objectiu realista: 100 converses iniciades, 10-15 propers a registre Premium quan arribi novembre."),
]

row = 7
for k in kpi_data:
    for col_idx, val in enumerate(k, start=2):
        cell = ws4.cell(row=row, column=col_idx, value=val)
        if col_idx == 2:
            style_data_cell(cell, wrap=True, size=10, bold=True, color=PRIMARY)
        elif col_idx == 3:
            style_data_cell(cell, wrap=True, size=9, color=HOVER)
            if val == "Sembrar":
                cell.fill = PatternFill("solid", fgColor=CREAM)
            elif val == "Llançar":
                cell.fill = PatternFill("solid", fgColor="EFE3D2")
            else:
                cell.fill = PatternFill("solid", fgColor="F0E8DC")
        elif col_idx in (4, 6, 8):  # objectius
            style_data_cell(cell, wrap=False, size=10, color=MUTED)
            cell.alignment = Alignment(horizontal="right", vertical="center")
        elif col_idx in (5, 7, 9):  # reals (Paolo omple)
            style_data_cell(cell, wrap=False, size=10, color=PRIMARY)
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.fill = PatternFill("solid", fgColor="FFF8E7")  # groc molt clar per indicar 'omple'm'
        else:  # notes
            style_data_cell(cell, wrap=True, size=9, color=MUTED)
    ws4.row_dimensions[row].height = 50
    row += 1

# Fila de total amb sumes
row_total = row
ws4.cell(row=row_total, column=2, value="TOTAL setembre")
style_data_cell(ws4.cell(row=row_total, column=2), size=11, bold=True, color=WHITE)
ws4.cell(row=row_total, column=2).fill = PatternFill("solid", fgColor=SECONDARY)
ws4.cell(row=row_total, column=2).alignment = Alignment(horizontal="left", vertical="center")
ws4.cell(row=row_total, column=3, value="—")
style_data_cell(ws4.cell(row=row_total, column=3), size=10, color=WHITE)
ws4.cell(row=row_total, column=3).fill = PatternFill("solid", fgColor=SECONDARY)
ws4.cell(row=row_total, column=3).alignment = Alignment(horizontal="center", vertical="center")
# Subscriptors newsletter final (última fila de setmana 5 setembre)
for col_idx, formula in [
    (4, f"=D{6+5}"),  # objectiu final = W5 set
    (5, f"=E{6+5}"),  # real final
    (6, f"=SUM(F7:F15)"),  # visitants acumulats
    (7, f"=SUM(G7:G15)"),
    (8, f"=H{6+5}"),  # seguidors final
    (9, f"=I{6+5}"),
]:
    cell = ws4.cell(row=row_total, column=col_idx, value=formula)
    style_data_cell(cell, size=11, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=SECONDARY)
    cell.alignment = Alignment(horizontal="right", vertical="center")
ws4.cell(row=row_total, column=10, value="—")
style_data_cell(ws4.cell(row=row_total, column=10), color=WHITE)
ws4.cell(row=row_total, column=10).fill = PatternFill("solid", fgColor=SECONDARY)
ws4.row_dimensions[row_total].height = 32

# Nota al peu
row_note = row_total + 2
ws4.cell(row=row_note, column=2, value="Llegenda: groc clar = camp per omplir cada diumenge. Objectius orientatius, no marxes forçades. Si els reals van per sobre, genial; si van per sota, és senyal d'ajustar el calendari, no de pitjar més.")
style_meta(ws4.cell(row=row_note, column=2))
ws4.merge_cells(start_row=row_note, start_column=2, end_row=row_note, end_column=10)
ws4.row_dimensions[row_note].height = 36
ws4.cell(row=row_note, column=2).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

ws4.freeze_panes = "C7"
ws4.sheet_view.zoomScale = 100

# ============================================================
# FULL 5 — GUIA DE TO
# ============================================================
ws5 = wb.create_sheet("Guia de to")
ws5.sheet_view.showGridLines = False

ws5['B2'] = "Guia de to · com escrivim"
style_title(ws5['B2'])
ws5.row_dimensions[2].height = 28
ws5['B3'] = "Criteri ESG té una veu. Aquesta veu és el nostre actiu. Sense ella, som un agregador més."
style_subtitle(ws5['B3'])

ws5['B4'] = ""
ws5['B4'].fill = PatternFill("solid", fgColor=ACCENT)
ws5.row_dimensions[4].height = 3
ws5.merge_cells('B4:E4')

# Bloc 1: 5 principis
ws5['B6'] = "Els 5 principis"
style_section_header(ws5['B6'])
ws5.row_dimensions[6].height = 24

principis_full = [
    ("1. Editorial, no comercial",
     "El nostre to és el d'un observatori que publica, no el d'una marca que ven. Si un text es pot confondre amb un anunci, cal reescriure'l."),
    ("2. Una dada per peça",
     "Cada peça té una sola idea amb una sola dada. La dada sempre és real, sempre està referenciada. Si calen dues dades, calen dues peces."),
    ("3. Sempre una pregunta",
     "Cada peça acaba amb una pregunta oberta, no amb un CTA. La pregunta crea relació; el CTA crea transacció. La relació és el nostre negoci."),
    ("4. Mai superlatius",
     "Prohibits: 'revolucionari', 'innovador', 'únic', 'el primer', 'el millor', 'líder', 'top'. Si ho som, que ho diguin els altres. Nosaltres només describim."),
    ("5. Referenciar sempre",
     "Cada afirmació sobre un informe o regulació porta la font. Sempre. És el nostre actiu de credibilitat en un món de greenwashing."),
]
row = 8
for label, txt in principis_full:
    ws5.cell(row=row, column=2, value=label)
    style_data_cell(ws5.cell(row=row, column=2), wrap=True, bold=True, color=HOVER, size=10)
    ws5.merge_cells(start_row=row, start_column=3, end_row=row, end_column=5)
    ws5.cell(row=row, column=3, value=txt)
    style_data_cell(ws5.cell(row=row, column=3), wrap=True, size=10)
    ws5.row_dimensions[row].height = 48
    row += 1

# Bloc 2: Frases prohibides vs alternatives
row += 2
ws5.cell(row=row, column=2, value="El que no diem / el que sí diem")
style_section_header(ws5.cell(row=row, column=2))
ws5.row_dimensions[row].height = 24
row += 1

headers5b = ["Tipus", "Frase prohibida", "Frase alternativa", "Per què"]
widths5b = [18, 38, 38, 38]
for i, (h, w) in enumerate(zip(headers5b, widths5b), start=2):
    col_letter = get_column_letter(i)
    ws5.column_dimensions[col_letter].width = w
    cell = ws5.cell(row=row, column=i, value=h)
    style_table_header(cell)
ws5.row_dimensions[row].height = 28
row += 1

frases = [
    ("Comercial", "Apúntate ya a la newsletter", "Si quieres recibir lo que publiquemos, apúntate.", "El 'ya' crea urgència artificial. La segona és una invitació."),
    ("Comercial", "No te lo pierdas", "Te lo compartimos por si te sirve.", "El primer és ordre. El segon és cura."),
    ("Comercial", "Oferta limitada: 50 plazas", "(no dir res fins a octubre)", "No venem. Quan arribi novembre, ja parlarem de preus."),
    ("Superlatiu", "La mejor plataforma ESG", "Una plataforma ESG (entre altres)", "Si som la millor, que ho digui un tercer. No nosaltres."),
    ("Superlatiu", "Único en el mercado", "Hemos encontrado un hueco que otros no cubren.", "Humilitat + descripció concreta."),
    ("Buit", "Solución integral ESG", "Sintetitzar informes, creuar amb certificacions, recomanar accions.", "La paraula integral no vol dir res. Els verbs sí."),
    ("Buit", "Valor añadido para tu negocio", "Estalviar 5 hores setmanals al teu equip.", "Concret, mesurable, verifiable."),
    ("Tòpic", "En un mundo cada vez más ESG...", "Hi ha 14 informes institucionals sobre ESG cada setmana.", "El tòpic no diu res. La dada sí."),
    ("Maniqueu", "Las empresas que no hagan ESG desaparecerán", "Las empresas que no procesen ESG perderán capacidad de decisión.", "No és el final del món, és una pèrdua de capacitat."),
    ("Maniqueu", "El futuro es sostenible o no es", "La pregunta no és si el futur és sostenible, és qui decidirà què vol dir sostenible.", "Matisos > eslògans."),
]
for f in frases:
    for col_idx, val in enumerate(f, start=2):
        cell = ws5.cell(row=row, column=col_idx, value=val)
        if col_idx == 2:
            style_data_cell(cell, wrap=True, size=9, bold=True, color=HOVER)
        elif col_idx == 3:
            style_data_cell(cell, wrap=True, size=10, color="A0522D")  # vermell terra per prohibit
            cell.font = Font(name=FONT_BODY, size=10, color="A0522D", strike=True)
        elif col_idx == 4:
            style_data_cell(cell, wrap=True, size=10, color="4A6B3A")  # verd terra per OK
        else:
            style_data_cell(cell, wrap=True, size=9, color=MUTED)
    ws5.row_dimensions[row].height = 36
    row += 1

# Bloc 3: Frases signatura
row += 2
ws5.cell(row=row, column=2, value="Frases signatura (es poden reciclar)")
style_section_header(ws5.cell(row=row, column=2))
ws5.row_dimensions[row].height = 24
row += 1

signatura = [
    "«La pregunta no és què publicarà la UE demà, és qui t'ajuda a saber què fer amb allò que ja s'ha publicat.»",
    "«El problema de l'ESG no és la manca d'informació. És la manca de temps per processar-la i de criteri per decidir què importa.»",
    "«La dada sense el teu context és soroll. La dada amb el teu context és decisió.»",
    "«No som un agregador. Som un filtre amb criteri.»",
    "«Cada quinzena, una connexió entre informes que no trobareu enlloc més.»",
]
for s in signatura:
    ws5.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
    cell = ws5.cell(row=row, column=2, value=s)
    cell.font = Font(name=FONT_TITLE, size=11, italic=True, color=PRIMARY)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws5.row_dimensions[row].height = 32
    row += 1

# Bloc 4: Formats per tipus de peça
row += 2
ws5.cell(row=row, column=2, value="Formats segons tipus de peça")
style_section_header(ws5.cell(row=row, column=2))
ws5.row_dimensions[row].height = 24
row += 1

headers5c = ["Tipus", "Estructura", "Extensió"]
widths5c = [22, 60, 22]
for i, (h, w) in enumerate(zip(headers5c, widths5c), start=2):
    col_letter = get_column_letter(i)
    ws5.column_dimensions[col_letter].width = w
    cell = ws5.cell(row=row, column=i, value=h)
    style_table_header(cell)
ws5.row_dimensions[row].height = 28
row += 1

formats_data = [
    ("Post curt LinkedIn", "1 idea + 1 dada + 1 pregunta. Sense CTA explícit. Salt de línia generós. Màxim 8 línies.", "60-100 paraules"),
    ("Article web/blog", "Tesi + dada + 2-3 casos + antítesi + pregunta final. Cap CTA fins al final, i subtil.", "700-1.500 paraules"),
    ("Peça visual LinkedIn", "Imatge 1080x1350 amb una frase curta (max 12 paraules). Text a la descripció: 3-4 línies + pregunta.", "Imatge + 30-50 paraules"),
    ("Newsletter", "Header + carta del director + 3 informes + connexió + notícies + inversió ESG + pregunta ètica. Sense CTAs comercials.", "1.500-2.000 paraules"),
    ("Email a la waitlist", "Salutació + una sola idea + una sola dada + què rebran la propera vegada. Cap més de 4 paràgrafs.", "200-300 paraules"),
]
for f in formats_data:
    for col_idx, val in enumerate(f, start=2):
        cell = ws5.cell(row=row, column=col_idx, value=val)
        if col_idx == 2:
            style_data_cell(cell, wrap=True, size=10, bold=True, color=PRIMARY)
        else:
            style_data_cell(cell, wrap=True, size=10)
        cell.fill = PatternFill("solid", fgColor=CREAM if row % 2 else WHITE)
    ws5.row_dimensions[row].height = 50
    row += 1

ws5.sheet_view.zoomScale = 100

# ============================================================
# FULL 6 — ASSETS VISUALS A PRODUIR
# ============================================================
ws6 = wb.create_sheet("Assets visuals")
ws6.sheet_view.showGridLines = False

ws6['B2'] = "Assets visuals a produir"
style_title(ws6['B2'])
ws6.row_dimensions[2].height = 28
ws6['B3'] = "Llista dels 14 assets visuals necessaris per executar el calendari. Tots segueixen la paleta terra+coure i tipografia Fraunces/Inter/JetBrains Mono."
style_subtitle(ws6['B3'])

ws6['B4'] = ""
ws6['B4'].fill = PatternFill("solid", fgColor=ACCENT)
ws6.row_dimensions[4].height = 3
ws6.merge_cells('B4:K4')

headers6 = ["ID", "Peça associada", "Descripció visual", "Format", "Dimensions", "Paleta", "Tipografia", "Responsable", "Data límit", "Estat"]
widths6 = [6, 26, 50, 12, 14, 28, 24, 14, 14, 12]
for i, (h, w) in enumerate(zip(headers6, widths6), start=2):
    col_letter = get_column_letter(i)
    ws6.column_dimensions[col_letter].width = w
    cell = ws6.cell(row=6, column=i, value=h)
    style_table_header(cell)
ws6.row_dimensions[6].height = 32

assets = [
    ("A1", "W1·agost post 'El soroll ESG'",
     "Imatge minimalista sobre fons crema #F5EFE6 amb la frase '14 informes. 0 tiempo.' centrada en tipografia Fraunces, mida gran, color marró #2C1810. Sense logo, sense icones. Només text i espai en blanc.",
     "PNG", "1080x1080",
     "Fons #F5EFE6, text #2C1810",
     "Fraunces 600",
     "Roser", "1 agost", "Pendent"),

    ("A2", "W1·agost post 'Què fa un director'",
     "Foto en blanc i negre d'una taula desordenada amb papers i un portàtil, o il·lustració minimalista amb 4 tasques superposades en format stack. To melancòlic, no estressant.",
     "PNG", "1080x1080",
     "B/N o sepia amb un toc #B87333",
     "(no aplica)",
     "Roser", "5 agost", "Pendent"),

    ("A3", "W2·agost mapa 16 estàndards (versió quadrada)",
     "Mapa visual amb els 16 estàndards organitzats en 3 columnes (Regulacions / Frameworks / Certificacions). Cada estàndard com a card amb franja esquerra 8px del color de la categoria. Fons crema #F5EFE6. Llegenda amb els 3 colors a la part inferior.",
     "PNG", "1080x1080",
     "Regulacions #5C3A1E · Frameworks #B87333 · Certificacions #E8C99A",
     "Fraunces (títols) + Inter (descripcions) + JetBrains Mono (meta)",
     "Roser + Z.ai", "10 agost", "Pendent"),

    ("A4", "W2·agost mapa 16 estàndards (versió horitzontal Open Graph)",
     "Mateixa disposició que A3 però adaptada a 1200x630 per a previews a LinkedIn i Twitter. Menys text, més jerarquia visual.",
     "PNG", "1200x630",
     "Igual que A3",
     "Igual que A3",
     "Roser", "10 agost", "Pendent"),

    ("A5", "W2·agost mapa 16 estàndards (PDF descarregable)",
     "Versió PDF del mapa amb una pàgina per estàndard: icona, descripció extensa, nombre d'informes cross-ref, accés. Capçalera de marca i peu amb URL.",
     "PDF", "A4 vertical",
     "Igual que A3",
     "Igual que A3",
     "Z.ai (text) + Roser (disseny)", "11 agost", "Pendent"),

    ("A6", "W2·agost post 'CSRD no és certificació'",
     "Detalle del mapa A3 amb CSRD destacat (potser un circle o border més gruixut) i text al costat: 'CSRD = Regulación'. Fons crema.",
     "PNG", "1080x1080",
     "Igual que A3 + accent #B87333 per al destacat",
     "Fraunces + Inter",
     "Roser", "13 agost", "Pendent"),

    ("A7", "W3·agost llista 30 fonts (visual)",
     "Taula visual amb 30 noms de fonts organitzades per tipologia (Reguladors / ONGs / Ratings / Acadèmia). Cada font amb logo o monograma. Fons crema, títol en Fraunces.",
     "PNG", "1080x1350",
     "Fons #F5EFE6 · accent #B87333",
     "Fraunces (títol) + Inter (llista)",
     "Z.ai + Roser", "17 agost", "Pendent"),

    ("A8", "W3·agost llista 30 fonts (PDF descarregable)",
     "Versió PDF de la llista amb una pàgina per font: nom, URL, cadència, per què importa, què en traiem. Capçalera de marca.",
     "PDF", "A4 vertical",
     "Igual que A7",
     "Igual que A7",
     "Z.ai (text) + Roser (disseny)", "18 agost", "Pendent"),

    ("A9", "W3·agost post 'El silenci dels que ja ho sabien'",
     "Imatge minimalista amb la frase 'saber sin poder demostrar' centrada sobre fons marró fosc #5C3A1E en color crema #F5EFE6. Sense res més.",
     "PNG", "1080x1350",
     "Fons #5C3A1E · text #F5EFE6",
     "Fraunces 600 italic",
     "Roser", "20 agost", "Pendent"),

    ("A10", "W4·agost Carta ètica #1 'Més enllà del Checkbox'",
     "Tipografia Fraunces sobre fons marró molt fosc #2C1810 amb la pregunta 'Si tu empresa desapareciera mañana, ¿quién lo notaría de verdad?' en color crema #F5EFE6. Format vertical.",
     "PNG", "1080x1350",
     "Fons #2C1810 · text #F5EFE6 · accent #D9A574",
     "Fraunces 600",
     "Roser", "24 agost", "Pendent"),

    ("A11", "W4·agost countdown '5 setmanes'",
     "Captura de pantalla de la web (hero o secció de landing) amb overlay subtil '5 setmanes' a una cantonada. Sense grans tipografies, sense urgència.",
     "PNG", "1200x630",
     "Captura web + overlay #B87333",
     "Inter",
     "Roser", "27 agost", "Pendent"),

    ("A12", "W2·setembre gràfic 'On es perd el temps'",
     "Gràfic de barres horitzontals amb 4 categories: Recopilar 60%, Llegir 15%, Analitzar 15%, Decidir 10%. Colors terra: #5C3A1E, #8A5526, #B87333, #E8C99A. Etiquetes clares, sense icones.",
     "PNG", "1200x630",
     "4 tonalitats terra",
     "Inter (labels) + Fraunces (títol)",
     "Z.ai (dades) + Roser (disseny)", "7 setembre", "Pendent"),

    ("A13", "W3·setembre Carta ètica #2",
     "Mateix format que A10 amb la pregunta #2: 'Si mañana te pidieran demostrar el impacto positivo de tu empresa con un solo dato, ¿cuál elegirías y por qué?'",
     "PNG", "1080x1350",
     "Igual que A10",
     "Fraunces 600",
     "Roser", "14 setembre", "Pendent"),

    ("A14", "W4·setembre Semàfor metodològic (visual gran)",
     "Visual del semàfor amb 5 dimensions (Scope 3, Plazos, Fuentes, Granularidad, Verificación) i nota final (A-D). Format explicatiu. Una flecha o esquema visual clar.",
     "PNG", "1200x630",
     "Verd #5C8A5C, groc #C9A961, vermell terra #A0522D + neutres",
     "Inter (labels) + Fraunces (nota) + JetBrains Mono (meta)",
     "Z.ai (text) + Roser (disseny)", "21 setembre", "Pendent"),
]

row = 7
for a in assets:
    for col_idx, val in enumerate(a, start=2):
        cell = ws6.cell(row=row, column=col_idx, value=val)
        if col_idx == 2:
            style_data_cell(cell, wrap=False, size=10, bold=True, color=PRIMARY)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif col_idx == 3:
            style_data_cell(cell, wrap=True, size=10, bold=True, color=PRIMARY)
        elif col_idx == 10:  # estat
            style_data_cell(cell, wrap=False, size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if val == "Pendent":
                cell.fill = PatternFill("solid", fgColor="FFE8B0")
                cell.font = Font(name=FONT_BODY, size=10, bold=True, color=HOVER)
            elif val == "En procés":
                cell.fill = PatternFill("solid", fgColor="E8C99A")
                cell.font = Font(name=FONT_BODY, size=10, bold=True, color=PRIMARY)
            elif val == "Fet":
                cell.fill = PatternFill("solid", fgColor="D8E8D0")
                cell.font = Font(name=FONT_BODY, size=10, bold=True, color="4A6B3A")
        else:
            style_data_cell(cell, wrap=True, size=9)
    ws6.row_dimensions[row].height = 60
    row += 1

# Conditional formatting per Estado
estado_range = f"K7:K{row-1}"
ws6.conditional_formatting.add(estado_range,
    CellIsRule(operator="equal", formula=['"Fet"'], fill=PatternFill("solid", fgColor="D8E8D0")))
ws6.conditional_formatting.add(estado_range,
    CellIsRule(operator="equal", formula=['"En procés"'], fill=PatternFill("solid", fgColor="E8C99A")))

ws6.freeze_panes = "C7"
ws6.sheet_view.zoomScale = 100

# Elimina el full per defecte
wb.remove(default)

# Guarda
output_path = "/home/z/my-project/download/pla-comunicacio-criteri-esg-agost-setembre-2026.xlsx"
wb.save(output_path)
print(f"✓ Excel guardat a {output_path}")
print(f"  Fulls: {wb.sheetnames}")
print(f"  Peces al calendari: {len(peces)}")
print(f"  Peces aliades: {len(aliades_data)}")
print(f"  Setmanes KPIs: {len(kpi_data)}")
print(f"  Assets visuals: {len(assets)}")
