import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "URLs a corregir"

# Column headers
headers = ["Font que dona problema", "URL actual (amb problemes)", "URL nova verificada"]
ws.append(headers)

# Style headers
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for col_idx, cell in enumerate(ws[1], 1):
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', wrap_text=True)
    cell.border = thin_border

# Dades: fonts que han fallat al test (basat en l'execució real)
problems = [
    # Reguladors europeus
    ("EFRAG ESRS Q&A Platform", "https://www.efrag.org/sustainability-reporting/esrs-q-a-platform", ""),
    ("EFRAG ESRS Knowledge Hub", "https://www.efrag.org/sustainability-reporting/esrs-knowledge-hub", ""),
    ("Comissió Europea DG CLIMA", "https://climate.ec.europa.eu/news-your-agenda/news_en", ""),
    ("European Environment Agency (EEA)", "https://www.eea.europa.eu/en/newsroom", ""),
    ("European Court of Auditors (ECA)", "https://www.eca.europa.eu/en/publications", ""),
    
    # Frameworks i estàndards
    ("TNFD", "https://tnfd.global/news-and-insights/", ""),
    ("SASB", "https://sasb.org/newsroom/", ""),
    ("TCFD (FSB)", "https://www.fsb.org/work-of-the-fsb/financial-innovation-and-structural-change/climate-related-financial-disclosures/", ""),
    ("EUR-Lex", "https://eur-lex.europa.eu/", ""),
    
    # Ratings i certificacions
    ("MSCI ESG Ratings", "https://www.msci.com/our-solutions/esg-investing", ""),
    ("ISS ESG", "https://www.issgovernance.com/esg/", ""),
    ("FTSE4Good", "https://www.ftserussell.com/products/indices/ftse4good", ""),
    ("Morningstar Sustainability", "https://www.morningstar.com/sustainability", ""),
    
    # Fonts globals
    ("World Economic Forum", "https://www.weforum.org/press/", ""),
    ("UNEP FI", "https://www.unepfi.org/wordpress/news/", ""),
    ("UNEP", "https://www.unep.org/news-and-stories", ""),
    ("OCDE", "https://www.oecd.org/newsroom/", ""),
    ("World Bank", "https://www.worldbank.org/en/news", ""),
    ("IMF Climate Change", "https://www.imf.org/en/Topics/climate-change", ""),
    ("IEA", "https://www.iea.org/news", ""),
    ("IRENA", "https://www.irena.org/News", ""),
    ("BIS", "https://www.bis.org/list/index.htm", ""),
    ("UNCTAD", "https://unctad.org/news", ""),
    
    # Think tanks i ONGs
    ("Carbon Tracker Initiative", "https://carbontracker.org/news/", ""),
    ("InfluenceMap", "https://influencemap.org/news", ""),
    ("Oxfam International", "https://www.oxfam.org/en/press-releases", ""),
    ("NewClimate Institute", "https://newclimate.org/resources/", ""),
    ("WWF", "https://www.worldwildlife.org/press", ""),
    ("Bruegel", "https://www.bruegel.org/analysis", ""),
    ("Transport & Environment", "https://www.transportenvironment.org/articles", ""),
    ("Agora Energiewende", "https://www.agora-energiewende.de/en/publications/", ""),
    
    # Fonts espanyoles
    ("CNMV", "https://www.cnmv.es/ES/Prensa", ""),
    ("MITECO", "https://www.miteco.gob.es/es/prensa/", ""),
    ("Banco de España", "https://www.bde.es/wbe/es/publicaciones/", ""),
    ("Forética", "https://foretica.org/noticias/", ""),
    ("CEOE", "https://www.ceoe.es/es/sala-de-prensa", ""),
    ("CES", "https://www.ces.es/web/guest/noticias", ""),
    ("Pimec", "https://www.pimec.org/es/actualidad", ""),
    ("Pacto Mundial ONU España", "https://www.pactomundial.org/noticias/", ""),
    
    # Fonts catalanes
    ("Departament Acció Climàtica (Gencat)", "https://mediambient.gencat.cat/ca/06_ambits_dactuacio/educacio_i_sostenibilitat/sala_de_premsa/", ""),
    ("CADS", "https://cads.gencat.cat/ca/inici/", ""),
    ("Oficina Catalana Canvi Climàtic", "https://canviclimatic.gencat.cat/ca/inici/", ""),
    ("Idescat", "https://www.idescat.cat/serveis/noticies/", ""),
    ("ACCIÓ", "https://accio.gencat.cat/ca/actualitat/", ""),
    ("Cambra de Comerç de Barcelona", "https://www.cambrabcn.org/ca/noticies", ""),
    
    # Consultories
    ("Deloitte CSRD insights", "https://dart.deloitte.com/USDW/home/news", ""),
    ("KPMG Sustainability Reporting", "https://kpmg.com/xx/en/what-we-do/industries/financial-services/sustainability-reporting.html", ""),
    ("EY Climate Action", "https://www.ey.com/en_gl/newsroom", ""),
    ("PwC Sustainability", "https://www.pwc.com/gx/en/news-room.html", ""),
    
    # Iniciatives
    ("Net Zero Tracker", "https://zerotracker.net/", ""),
    ("WBCSD", "https://www.wbcsd.org/news/", ""),
]

for row in problems:
    ws.append(list(row))

# Apply borders and formatting
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=3):
    for cell in row:
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True, vertical='top')

# Set column widths
ws.column_dimensions['A'].width = 45
ws.column_dimensions['B'].width = 70
ws.column_dimensions['C'].width = 70

# Freeze header row
ws.freeze_panes = 'A2'

# Save
output_path = r"C:\Users\dvd_f\criteri-esg-lab\scripts\scraper_urls_a_corregir.xlsx"
wb.save(output_path)
print(f"Excel guardat a: {output_path}")
print(f"Total fonts a revisar: {len(problems)}")